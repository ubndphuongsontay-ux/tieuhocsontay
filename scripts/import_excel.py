"""Nhập GV + HS năm 2026-2027 vào Postgres local th_son_tay."""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import psycopg
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
REPORT_PATH = ROOT / "scripts" / "import_report.json"

GV_XLSX = Path(r"C:\Users\cuong\Downloads\Danh_sach_giao vien.xlsx")
HS_XLSX = Path(r"C:\Users\cuong\Downloads\Danh_sach_hoc_sinh_chuan_2026-2027.xlsx")
PT_DIR = Path(r"C:\Users\cuong\Downloads\DANH SACH LOP")

DSN = "postgresql://postgres@127.0.0.1:54329/th_son_tay"
YEAR_START = date(2026, 9, 5)

CAMPUS_BY_NAME = {
    "trung hưng": "TH",
    "phú thịnh": "PT",
    "trần phú": "TP",
    "đường lâm": "DL",
    "lê lợi": "LL",
    "quang trung": "QT",
    "viên sơn": "VS",
}
LETTER_TO_CAMPUS = {"A": "TH", "C": "PT", "D": "TP", "E": "DL", "G": "LL", "H": "QT"}


def nonempty(v) -> bool:
    if v is None:
        return False
    if isinstance(v, float) and pd.isna(v):
        return False
    s = str(v).strip()
    return s != "" and s.lower() not in {"nan", "none", "nat"}


def text(v) -> str | None:
    if not nonempty(v):
        return None
    return str(v).strip()


def parse_date(v) -> date | None:
    if not nonempty(v):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    s = s.replace(".", "/").replace("-", "/")
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:19] if ":" in s else s, fmt).date()
        except ValueError:
            continue
    try:
        n = float(str(v))
        if n > 20000:
            return (datetime(1899, 12, 30) + pd.to_timedelta(n, unit="D")).date()
    except (TypeError, ValueError):
        pass
    return None


def gender_of(v) -> str | None:
    s = text(v)
    if not s:
        return None
    t = s.lower()
    if t.startswith("nam") and "nữ" not in t and "nu" not in t:
        return "nam"
    if t in {"nam"}:
        return "nam"
    if "nữ" in t or t in {"nu", "nữ"}:
        return "nu"
    if t == "nữ":
        return "nu"
    return None


def party_of(v) -> bool | None:
    s = text(v)
    if not s:
        return None
    t = s.lower()
    if t in {"có", "co", "true", "1", "x"}:
        return True
    if t in {"không", "khong", "false", "0"}:
        return False
    return None


def split_national_id(raw) -> tuple[str | None, str | None]:
    s = text(raw)
    if not s:
        return None, None
    digits = re.sub(r"\D", "", s)
    nid = digits if len(digits) == 12 else None
    return nid, s


def campus_code_from_name(name: str | None) -> str | None:
    if not name:
        return None
    key = re.sub(r"\s+", " ", name.strip().lower())
    if key in CAMPUS_BY_NAME:
        return CAMPUS_BY_NAME[key]
    for k, code in CAMPUS_BY_NAME.items():
        if k in key:
            return code
    return None


def class_meta(class_name: str | None, campus_hint: str | None) -> tuple[str, int, str]:
    name = (class_name or "").strip().upper().replace(" ", "")
    m = re.match(r"^(\d)([A-Z])(\d+)$", name)
    if not m:
        raise ValueError(f"Tên lớp không nhận dạng được: {class_name!r}")
    grade = int(m.group(1))
    letter = m.group(2)
    code = campus_hint or LETTER_TO_CAMPUS.get(letter)
    if not code:
        raise ValueError(f"Không suy được phân hiệu từ lớp {name}")
    return name, grade, code


def is_header_row(row) -> bool:
    vals = [str(c).strip().lower() if c is not None else "" for c in row]
    return "stt" in vals and any("họ tên" in v or "ho ten" in v for v in vals)


def main() -> int:
    report: dict = {"warnings": [], "issues": []}

    with psycopg.connect(DSN) as conn:
        conn.execute("set client_encoding to 'UTF8'")
        campuses = {row[0]: row[1] for row in conn.execute("select code, id from campuses")}
        year = conn.execute(
            "select id, starts_on from school_years where code = %s",
            ("2026-2027",),
        ).fetchone()
        if not year:
            raise SystemExit("Chưa có năm học 2026-2027")
        year_id = year[0]
        started_on = year[1] or YEAR_START

        n_staff = import_staff(conn, campuses, year_id, report)
        n_hs, n_enr = import_students(conn, campuses, year_id, started_on, report)
        conn.commit()

    with psycopg.connect(DSN) as conn:
        stats = {
            "schools": conn.execute("select count(*) from schools").fetchone()[0],
            "campuses": conn.execute("select count(*) from campuses").fetchone()[0],
            "staff": conn.execute("select count(*) from staff").fetchone()[0],
            "staff_reviews": conn.execute("select count(*) from staff_reviews").fetchone()[0],
            "students": conn.execute("select count(*) from students").fetchone()[0],
            "classes": conn.execute("select count(*) from classes").fetchone()[0],
            "enrollments": conn.execute("select count(*) from enrollments").fetchone()[0],
            "enrollments_open": conn.execute(
                "select count(*) from enrollments where ended_on is null"
            ).fetchone()[0],
            "contacts": conn.execute("select count(*) from student_contacts").fetchone()[0],
            "supports": conn.execute("select count(*) from student_supports").fetchone()[0],
        }
        by_campus = [
            {"campus_code": r[0], "campus_name": r[1], "hs": r[2]}
            for r in conn.execute(
                """
                select campus_code, campus_name, count(*)::int as hs
                from v_enrollments_current
                group by campus_code, campus_name
                order by campus_code
                """
            )
        ]
        empty_vs = conn.execute(
            "select count(*) from v_enrollments_current where campus_code = 'VS'"
        ).fetchone()[0]

    report["imported"] = {
        "staff_rows": n_staff,
        "student_rows": n_hs,
        "enrollment_rows": n_enr,
    }
    report["db"] = stats
    report["hs_by_campus"] = by_campus
    report["vien_son_open"] = empty_vs
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["db"], ensure_ascii=False, indent=2))
    print("HS theo phân hiệu:", by_campus)
    print("Cảnh báo:", len(report["warnings"]), "vấn đề:", len(report["issues"]))
    print("Báo cáo:", REPORT_PATH)
    return 0


def import_staff(conn, campuses: dict, year_id, report: dict) -> int:
    df = pd.read_excel(GV_XLSX, sheet_name="Dữ liệu chuẩn hóa", dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    n = 0
    for _, row in df.iterrows():
        code = campus_code_from_name(text(row.get("Phân hiệu")))
        if not code:
            report["issues"].append({"kind": "gv_no_campus", "name": text(row.get("Họ tên"))})
            continue
        nid, raw = split_national_id(row.get("Mã số định danh cá nhân"))
        rating = text(row.get("Đánh giá viên chức"))
        employment = None
        if rating and "thỉnh" in rating.lower():
            employment = "thinh_giang"
            rating = None
        staff_id = conn.execute(
            """
            insert into staff (
              campus_id, full_name, national_id, national_id_raw, dob, gender,
              ethnicity, phone, is_party_member, education_level,
              foreign_language_level, it_level, political_theory_level,
              professional_qualification, employment_kind
            ) values (
              %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
            returning id
            """,
            (
                campuses[code],
                text(row.get("Họ tên")),
                nid,
                raw,
                parse_date(row.get("Ngày sinh")),
                gender_of(row.get("Giới tính")),
                text(row.get("Dân tộc")),
                text(row.get("Điện thoại")),
                party_of(row.get("Đảng viên")),
                text(row.get("Trình độ chính")),
                text(row.get("T.độ Đ.tạo N.Ngữ")),
                text(row.get("T.Độ tin học")),
                text(row.get("T.Độ lý luận chính trị")),
                text(row.get("T.Độ chuyên môn nghiệp vụ")),
                employment,
            ),
        ).fetchone()[0]
        if rating:
            conn.execute(
                """
                insert into staff_reviews (staff_id, school_year_id, rating)
                values (%s, %s, %s)
                """,
                (staff_id, year_id, rating),
            )
        n += 1
    return n


def upsert_class(conn, cache: dict, year_id, campus_id, name: str, grade: int):
    key = (str(campus_id), name)
    if key in cache:
        return cache[key]
    cid = conn.execute(
        """
        insert into classes (school_year_id, campus_id, name, grade)
        values (%s, %s, %s, %s)
        on conflict (school_year_id, campus_id, name) do update
          set grade = excluded.grade
        returning id
        """,
        (year_id, campus_id, name, grade),
    ).fetchone()[0]
    cache[key] = cid
    return cid


def find_or_create_student(conn, rec: dict, nid_owner: dict, report: dict):
    nid = rec["national_id"]
    bgd = rec["bgd_code"]
    if nid and nid in nid_owner:
        existing = nid_owner[nid]
        same = (
            existing["full_name"] == rec["full_name"]
            and existing["dob"] == rec["dob"]
        ) or existing["full_name"].split("(")[0].strip() == rec["full_name"].split("(")[0].strip()
        if same or existing["full_name"] == rec["full_name"]:
            return existing["id"], True
        rec["national_id"] = None
        report["issues"].append(
            {
                "kind": "cccd_trung_khac_nguoi",
                "national_id_raw": rec["national_id_raw"],
                "name": rec["full_name"],
                "other": existing["full_name"],
            }
        )
        nid = None
    if bgd:
        found = conn.execute(
            "select id from students where bgd_code = %s", (bgd,)
        ).fetchone()
        if found:
            return found[0], True
    sid = conn.execute(
        """
        insert into students (
          full_name, dob, gender, ethnicity, national_id, national_id_raw, bgd_code
        ) values (%s,%s,%s,%s,%s,%s,%s)
        returning id
        """,
        (
            rec["full_name"],
            rec["dob"],
            rec["gender"],
            rec["ethnicity"],
            nid,
            rec["national_id_raw"],
            bgd,
        ),
    ).fetchone()[0]
    if nid:
        nid_owner[nid] = {
            "id": sid,
            "full_name": rec["full_name"],
            "dob": rec["dob"],
        }
    return sid, False


def close_open_enrollment(conn, student_id, year_id, new_campus_id, report, name, class_name):
    open_row = conn.execute(
        """
        select e.id, c.campus_id, c.name
        from enrollments e
        join classes c on c.id = e.class_id
        where e.student_id = %s and e.school_year_id = %s and e.ended_on is null
        """,
        (student_id, year_id),
    ).fetchone()
    if not open_row:
        return
    old_id, old_campus, old_class = open_row
    status = "chuyen_phan_hieu" if old_campus != new_campus_id else "chuyen_lop"
    if "chuyển trường" in (name or "").lower() or "chuyen truong" in (name or "").lower():
        status = "chuyen_truong"
    conn.execute(
        """
        update enrollments
        set ended_on = %s, status = %s
        where id = %s
        """,
        (YEAR_START, status, old_id),
    )
    report["warnings"].append(
        {
            "kind": "dong_cho_hoc_cu",
            "name": name,
            "from": old_class,
            "to": class_name,
            "status": status,
        }
    )


def add_contacts_and_supports(conn, student_id, year_id, rec: dict):
    contacts = []
    if rec.get("mom") or rec.get("mom_phone"):
        contacts.append(("me", rec.get("mom"), rec.get("mom_phone")))
    if rec.get("dad") or rec.get("dad_phone"):
        contacts.append(("cha", rec.get("dad"), rec.get("dad_phone")))
    if rec.get("other_phone"):
        contacts.append(("khac", None, rec.get("other_phone")))
    primary_rel = "khac" if rec.get("other_phone") else ("me" if rec.get("mom") or rec.get("mom_phone") else "cha")
    for rel, nm, phone in contacts:
        conn.execute(
            """
            insert into student_contacts (student_id, relation, full_name, phone, is_primary)
            values (%s,%s,%s,%s,%s)
            """,
            (student_id, rel, nm, phone, rel == primary_rel),
        )
    if rec.get("disability"):
        conn.execute(
            """
            insert into student_supports (student_id, school_year_id, kind, label, opened_on)
            values (%s,%s,'khuyet_tat',%s,%s)
            """,
            (student_id, year_id, rec["disability"], YEAR_START),
        )
    if rec.get("policy"):
        conn.execute(
            """
            insert into student_supports (student_id, school_year_id, kind, label, opened_on)
            values (%s,%s,'chinh_sach',%s,%s)
            """,
            (student_id, year_id, rec["policy"], YEAR_START),
        )
    if rec.get("doi_tuong"):
        conn.execute(
            """
            insert into student_supports (student_id, school_year_id, kind, label, opened_on)
            values (%s,%s,'doi_tuong',%s,%s)
            """,
            (student_id, year_id, rec["doi_tuong"], YEAR_START),
        )


def enroll(conn, student_id, class_id, year_id, name, campus_id, class_name, is_existing, report):
    if is_existing:
        close_open_enrollment(conn, student_id, year_id, campus_id, report, name, class_name)
    conn.execute(
        """
        insert into enrollments (
          student_id, class_id, school_year_id, started_on, ended_on, status
        ) values (%s,%s,%s,%s,null,'dang_hoc')
        """,
        (student_id, class_id, year_id, YEAR_START),
    )


def import_students(conn, campuses, year_id, started_on, report) -> tuple[int, int]:
    class_cache: dict = {}
    nid_owner: dict = {}
    n_hs = 0
    n_enr = 0
    seen_student_ids: set = set()

    df = pd.read_excel(HS_XLSX, sheet_name="Danh_sach_chuan", dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    for _, row in df.iterrows():
        name = text(row.get("Họ tên"))
        if not name:
            continue
        campus_code = campus_code_from_name(text(row.get("Điểm/Phân hiệu")))
        cls_name, grade, inferred = class_meta(text(row.get("Lớp")), campus_code)
        code = campus_code or inferred
        nid, raw = split_national_id(row.get("Số định danh cá nhân"))
        rec = {
            "full_name": name,
            "dob": parse_date(row.get("Ngày sinh")),
            "gender": gender_of(row.get("Giới tính")),
            "ethnicity": text(row.get("Dân tộc")),
            "national_id": nid,
            "national_id_raw": raw,
            "bgd_code": text(row.get("Mã định danh Bộ GD&ĐT")),
            "mom": text(row.get("Tên mẹ")),
            "mom_phone": text(row.get("SĐT mẹ")),
            "dad": text(row.get("Tên cha")),
            "dad_phone": text(row.get("SĐT cha")),
            "other_phone": text(row.get("SĐT liên hệ")),
            "disability": text(row.get("Loại khuyết tật")),
            "policy": text(row.get("Đối tượng chính sách")),
            "doi_tuong": text(row.get("Đối tượng HS")),
        }
        sid, existed = find_or_create_student(conn, rec, nid_owner, report)
        class_id = upsert_class(conn, class_cache, year_id, campuses[code], cls_name, grade)
        enroll(conn, sid, class_id, year_id, name, campuses[code], cls_name, existed, report)
        n_enr += 1
        if sid not in seen_student_ids:
            add_contacts_and_supports(conn, sid, year_id, rec)
            seen_student_ids.add(sid)
            n_hs += 1

    for f in sorted(PT_DIR.rglob("*.xlsx")):
        recs = parse_phu_thinh_file(f, report)
        for rec in recs:
            sid, existed = find_or_create_student(conn, rec, nid_owner, report)
            class_id = upsert_class(
                conn, class_cache, year_id, campuses["PT"], rec["class_name"], rec["grade"]
            )
            enroll(
                conn, sid, class_id, year_id, rec["full_name"], campuses["PT"], rec["class_name"], existed, report
            )
            n_enr += 1
            if sid not in seen_student_ids:
                add_contacts_and_supports(conn, sid, year_id, rec)
                seen_student_ids.add(sid)
                n_hs += 1
    return n_hs, n_enr


def parse_phu_thinh_file(path: Path, report: dict) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw = list(ws.iter_rows(max_col=12, values_only=True))
    wb.close()
    header_idx = None
    for i, row in enumerate(raw):
        if is_header_row(row):
            header_idx = i
            break
    if header_idx is None:
        report["issues"].append({"kind": "pt_no_header", "file": path.name})
        return []
    headers = [str(c).strip() if c is not None else f"c{j}" for j, c in enumerate(raw[header_idx])]
    title = " ".join(str(c) for r in raw[:6] for c in r if c)
    m = re.search(r"LỚP\s+(\dC\d)", title, re.I)
    class_name = m.group(1).upper() if m else None
    if not class_name:
        m2 = re.search(r"(\dC\d)", path.name, re.I)
        class_name = m2.group(1).upper() if m2 else None
    if not class_name:
        report["issues"].append({"kind": "pt_no_class", "file": path.name})
        return []
    _, grade, _ = class_meta(class_name, "PT")
    out = []
    for row in raw[header_idx + 1 :]:
        stt = row[0] if row else None
        try:
            int(float(str(stt).strip()))
        except (TypeError, ValueError):
            continue
        by = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        name = text(by.get("Họ tên") or by.get("Họ và tên"))
        if not name:
            continue
        nid, raw_id = split_national_id(by.get("Số định danh cá nhân"))
        out.append(
            {
                "full_name": name,
                "dob": parse_date(by.get("Ngày sinh")),
                "gender": gender_of(by.get("Giới tính")),
                "ethnicity": text(by.get("Dân tộc")),
                "national_id": nid,
                "national_id_raw": raw_id,
                "bgd_code": None,
                "mom": text(by.get("Tên mẹ")),
                "mom_phone": text(by.get("SĐT mẹ")),
                "dad": text(by.get("Tên cha")),
                "dad_phone": text(by.get("SĐT cha")),
                "other_phone": None,
                "disability": text(by.get("Loại khuyết tật")),
                "policy": text(by.get("Đối tượng chính sách")),
                "doi_tuong": None,
                "class_name": class_name,
                "grade": grade,
            }
        )
    return out


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print("IMPORT FAILED:", exc, file=sys.stderr)
        raise
